#!/usr/bin/env python3
"""Deterministic and constrained XLSX operations for ArcForge."""

from __future__ import annotations

import argparse
import ast
import hashlib
import json
import os
import re
import sys
import tempfile
from copy import copy
from datetime import date, datetime, timedelta
from decimal import Decimal
from pathlib import Path
from typing import Any, Dict, Iterable, List, Mapping, Optional, Sequence, Tuple

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.chart import AreaChart, BarChart, LineChart, PieChart, Reference
    from openpyxl.comments import Comment
    from openpyxl.styles import (
        Alignment,
        Border,
        Font,
        GradientFill,
        NamedStyle,
        PatternFill,
        Protection,
        Side,
    )
    from openpyxl.utils import get_column_letter, range_boundaries
    from openpyxl.worksheet.table import Table, TableStyleInfo

    OPENPYXL_IMPORT_ERROR: Optional[Exception] = None
except (
    Exception
) as error:  # pragma: no cover - exercised only when dependency is absent
    OPENPYXL_IMPORT_ERROR = error


DEFAULT_THEME = {
    "primary": "0F172A",
    "accent": "2563EB",
    "header_fill": "0F172A",
    "header_text": "FFFFFF",
}
MAX_INSPECTED_CELLS = 250_000
MAX_SCRIPT_BYTES = 128 * 1024
MAX_SCRIPT_AST_NODES = 20_000
MAX_RANGE_ITEMS = 1_000_000
TABLE_NAME_RE = re.compile(r"^[A-Za-z_][A-Za-z0-9_.]*$")

FORBIDDEN_CODE_NODES = (
    ast.AsyncFor,
    ast.AsyncFunctionDef,
    ast.AsyncWith,
    ast.Await,
    ast.ClassDef,
    ast.Global,
    ast.Import,
    ast.ImportFrom,
    ast.Nonlocal,
)
FORBIDDEN_CODE_NAMES = {
    "__builtins__",
    "__import__",
    "breakpoint",
    "compile",
    "delattr",
    "dir",
    "eval",
    "exec",
    "getattr",
    "globals",
    "help",
    "input",
    "locals",
    "open",
    "setattr",
    "vars",
}
FORBIDDEN_CODE_ATTRIBUTES = {
    "close",
    "save",
}
PROTECTED_CODE_NAMES = {
    "workbook",
    "Alignment",
    "AreaChart",
    "BarChart",
    "Border",
    "Decimal",
    "Font",
    "GradientFill",
    "LineChart",
    "NamedStyle",
    "PatternFill",
    "PieChart",
    "Protection",
    "Reference",
    "Side",
    "Table",
    "TableStyleInfo",
    "Comment",
    "copy",
    "date",
    "datetime",
    "get_column_letter",
    "timedelta",
}


class SpreadsheetError(RuntimeError):
    pass


class SpreadsheetCodeValidator(ast.NodeVisitor):
    """Reject syntax that could escape the workbook-only execution surface."""

    def __init__(self) -> None:
        self.node_count = 0

    @staticmethod
    def _location(node: ast.AST) -> str:
        line = getattr(node, "lineno", None)
        column = getattr(node, "col_offset", None)
        if line is None:
            return ""
        return f" at line {line}, column {(column or 0) + 1}"

    def _reject(self, node: ast.AST, message: str) -> None:
        raise SpreadsheetError(message + self._location(node))

    def generic_visit(self, node: ast.AST) -> None:
        self.node_count += 1
        if self.node_count > MAX_SCRIPT_AST_NODES:
            self._reject(
                node,
                f"Spreadsheet code exceeds the {MAX_SCRIPT_AST_NODES}-node AST limit",
            )
        if isinstance(node, FORBIDDEN_CODE_NODES):
            self._reject(
                node,
                f"Spreadsheet code cannot use {type(node).__name__}",
            )
        super().generic_visit(node)

    def visit_Name(self, node: ast.Name) -> None:
        if node.id.startswith("_") or node.id in FORBIDDEN_CODE_NAMES:
            self._reject(node, f"Spreadsheet code cannot access name {node.id!r}")
        if isinstance(node.ctx, (ast.Store, ast.Del)) and node.id in PROTECTED_CODE_NAMES:
            self._reject(node, f"Spreadsheet code cannot rebind protected name {node.id!r}")
        self.generic_visit(node)

    def visit_Attribute(self, node: ast.Attribute) -> None:
        if node.attr.startswith("_") or node.attr in FORBIDDEN_CODE_ATTRIBUTES:
            self._reject(node, f"Spreadsheet code cannot access attribute {node.attr!r}")
        self.generic_visit(node)


def safe_range(*arguments: int) -> range:
    try:
        value = range(*arguments)
    except (TypeError, ValueError) as error:
        raise SpreadsheetError("Invalid range() arguments: " + str(error)) from error
    if len(value) > MAX_RANGE_ITEMS:
        raise SpreadsheetError(
            f"range() is limited to {MAX_RANGE_ITEMS} items in SpreadsheetCode"
        )
    return value


SAFE_CODE_BUILTINS: Dict[str, Any] = {
    "Exception": Exception,
    "KeyError": KeyError,
    "RuntimeError": RuntimeError,
    "TypeError": TypeError,
    "ValueError": ValueError,
    "abs": abs,
    "all": all,
    "any": any,
    "bool": bool,
    "dict": dict,
    "enumerate": enumerate,
    "float": float,
    "int": int,
    "isinstance": isinstance,
    "len": len,
    "list": list,
    "max": max,
    "min": min,
    "range": safe_range,
    "reversed": reversed,
    "round": round,
    "set": set,
    "sorted": sorted,
    "str": str,
    "sum": sum,
    "tuple": tuple,
    "zip": zip,
}


def require_openpyxl() -> None:
    if OPENPYXL_IMPORT_ERROR is not None:
        raise SpreadsheetError(
            "openpyxl is required. Ask the user before installing "
            "scripts/requirements.txt. Import error: " + str(OPENPYXL_IMPORT_ERROR)
        )


def load_json_object(path_value: str) -> Tuple[Dict[str, Any], Path]:
    path = Path(path_value).expanduser().resolve()
    if not path.is_file():
        raise SpreadsheetError("JSON specification does not exist: " + str(path))
    try:
        with path.open("r", encoding="utf-8-sig") as handle:
            value = json.load(handle)
    except (OSError, json.JSONDecodeError) as error:
        raise SpreadsheetError(
            "Failed to read JSON specification: " + str(error)
        ) from error
    if not isinstance(value, dict):
        raise SpreadsheetError("JSON specification must be an object")
    return value, path


def positive_int(value: Any, field: str, default: int = 1) -> int:
    if value is None:
        return default
    if isinstance(value, bool):
        raise SpreadsheetError(field + " must be an integer")
    try:
        result = int(value)
    except (TypeError, ValueError) as error:
        raise SpreadsheetError(field + " must be an integer") from error
    if result < 1:
        raise SpreadsheetError(field + " must be at least 1")
    return result


def normalize_color(value: Any, field: str = "color") -> str:
    raw = str(value).strip().lstrip("#").upper()
    if len(raw) == 6:
        raw = "FF" + raw
    if len(raw) != 8 or any(char not in "0123456789ABCDEF" for char in raw):
        raise SpreadsheetError(field + " must be a 6- or 8-digit hexadecimal color")
    return raw


def merged_theme(spec: Mapping[str, Any]) -> Dict[str, str]:
    theme = dict(DEFAULT_THEME)
    raw = spec.get("theme")
    if raw is None:
        return theme
    if not isinstance(raw, dict):
        raise SpreadsheetError("theme must be an object")
    for key in DEFAULT_THEME:
        if key in raw:
            theme[key] = normalize_color(raw[key], "theme." + key)[-6:]
    return theme


def apply_style(cell: Any, style: Mapping[str, Any]) -> None:
    if not isinstance(style, dict):
        raise SpreadsheetError("style must be an object")

    font_spec = style.get("font")
    if font_spec is not None:
        if not isinstance(font_spec, dict):
            raise SpreadsheetError("style.font must be an object")
        font = copy(cell.font)
        for key in ("name", "bold", "italic", "underline"):
            if key in font_spec:
                setattr(font, key, font_spec[key])
        if "size" in font_spec:
            font.sz = float(font_spec["size"])
        if "color" in font_spec:
            font.color = normalize_color(font_spec["color"], "style.font.color")
        cell.font = font

    fill_spec = style.get("fill")
    if fill_spec is not None:
        if isinstance(fill_spec, str):
            color = normalize_color(fill_spec, "style.fill")
        elif isinstance(fill_spec, dict) and "color" in fill_spec:
            color = normalize_color(fill_spec["color"], "style.fill.color")
        else:
            raise SpreadsheetError(
                "style.fill must be a color string or an object with color"
            )
        cell.fill = PatternFill(fill_type="solid", fgColor=color)

    alignment_spec = style.get("alignment")
    if alignment_spec is not None:
        if not isinstance(alignment_spec, dict):
            raise SpreadsheetError("style.alignment must be an object")
        alignment = copy(cell.alignment)
        allowed = {
            "horizontal",
            "vertical",
            "wrap_text",
            "shrink_to_fit",
            "text_rotation",
        }
        for key, value in alignment_spec.items():
            if key not in allowed:
                raise SpreadsheetError("Unsupported alignment field: " + str(key))
            setattr(alignment, key, value)
        cell.alignment = alignment

    border_spec = style.get("border")
    if border_spec is not None:
        if not isinstance(border_spec, dict):
            raise SpreadsheetError("style.border must be an object")
        side = Side(
            style=str(border_spec.get("style", "thin")),
            color=normalize_color(
                border_spec.get("color", "CBD5E1"), "style.border.color"
            ),
        )
        cell.border = Border(left=side, right=side, top=side, bottom=side)

    if "number_format" in style:
        cell.number_format = str(style["number_format"])


def decode_cell_value(raw: Any) -> Tuple[Any, Optional[Mapping[str, Any]]]:
    if not isinstance(raw, dict):
        return raw, None

    if "formula" in raw:
        formula = str(raw["formula"])
        value: Any = formula if formula.startswith("=") else "=" + formula
    else:
        value = raw.get("value")

    value_type = str(raw.get("type", "")).lower()
    if value is not None and value_type == "date":
        value = date.fromisoformat(str(value))
    elif value is not None and value_type == "datetime":
        value = datetime.fromisoformat(str(value))
    return value, raw


def set_cell(cell: Any, raw: Any) -> None:
    value, descriptor = decode_cell_value(raw)
    cell.value = value
    if descriptor is None:
        return
    if "style" in descriptor:
        apply_style(cell, descriptor["style"])
    if "number_format" in descriptor:
        cell.number_format = str(descriptor["number_format"])
    if descriptor.get("hyperlink"):
        cell.hyperlink = str(descriptor["hyperlink"])
        font = copy(cell.font)
        font.color = normalize_color("2563EB")
        font.underline = "single"
        cell.font = font
    if descriptor.get("comment"):
        cell.comment = Comment(
            str(descriptor["comment"]), str(descriptor.get("author", "ArcForge"))
        )


def write_rows(
    worksheet: Any,
    rows: Sequence[Any],
    start_row: int = 1,
    start_column: int = 1,
) -> None:
    if not isinstance(rows, list):
        raise SpreadsheetError("rows must be an array")
    for row_offset, raw_row in enumerate(rows):
        if not isinstance(raw_row, list):
            raise SpreadsheetError("each rows entry must be an array")
        for column_offset, raw_value in enumerate(raw_row):
            set_cell(
                worksheet.cell(
                    row=start_row + row_offset,
                    column=start_column + column_offset,
                ),
                raw_value,
            )


def iter_cells_in_range(worksheet: Any, range_ref: str) -> Iterable[Any]:
    try:
        min_column, min_row, max_column, max_row = range_boundaries(range_ref)
    except ValueError as error:
        raise SpreadsheetError("Invalid cell range: " + range_ref) from error
    for row in worksheet.iter_rows(
        min_row=min_row,
        max_row=max_row,
        min_col=min_column,
        max_col=max_column,
    ):
        for cell in row:
            yield cell


def apply_cell_map(worksheet: Any, raw: Any) -> None:
    if raw is None:
        return
    if isinstance(raw, dict):
        items = [{"cell": key, "value": value} for key, value in raw.items()]
    elif isinstance(raw, list):
        items = raw
    else:
        raise SpreadsheetError("set_cells must be an object or array")
    for item in items:
        if not isinstance(item, dict) or not item.get("cell"):
            raise SpreadsheetError("each set_cells entry requires a cell")
        set_cell(worksheet[str(item["cell"])], item.get("value"))


def apply_dimensions(worksheet: Any, sheet_spec: Mapping[str, Any]) -> None:
    widths = sheet_spec.get("column_widths", {})
    if not isinstance(widths, dict):
        raise SpreadsheetError("column_widths must be an object")
    for raw_column, raw_width in widths.items():
        column = str(raw_column).strip().upper()
        if column.isdigit():
            column = get_column_letter(positive_int(column, "column_widths key"))
        if not column.isalpha():
            raise SpreadsheetError("Invalid column_widths key: " + str(raw_column))
        worksheet.column_dimensions[column].width = float(raw_width)

    heights = sheet_spec.get("row_heights", {})
    if not isinstance(heights, dict):
        raise SpreadsheetError("row_heights must be an object")
    for raw_row, raw_height in heights.items():
        row = positive_int(raw_row, "row_heights key")
        worksheet.row_dimensions[row].height = float(raw_height)


def apply_styles(
    worksheet: Any, sheet_spec: Mapping[str, Any], theme: Mapping[str, str]
) -> None:
    header_row = sheet_spec.get("header_row")
    if header_row is not None and worksheet.max_column > 0:
        row_number = positive_int(header_row, "header_row")
        header_style = {
            "font": {"bold": True, "color": theme["header_text"]},
            "fill": theme["header_fill"],
            "alignment": {"vertical": "center", "wrap_text": True},
            "border": {"style": "thin", "color": theme["primary"]},
        }
        for cell in worksheet[row_number]:
            apply_style(cell, header_style)
        worksheet.row_dimensions[row_number].height = max(
            worksheet.row_dimensions[row_number].height or 0,
            24,
        )

    styles = sheet_spec.get("styles", [])
    if not isinstance(styles, list):
        raise SpreadsheetError("styles must be an array")
    for item in styles:
        if (
            not isinstance(item, dict)
            or not item.get("range")
            or not isinstance(item.get("style"), dict)
        ):
            raise SpreadsheetError("each styles entry requires range and style")
        for cell in iter_cells_in_range(worksheet, str(item["range"])):
            apply_style(cell, item["style"])


def reference_from_range(worksheet: Any, range_ref: str) -> Any:
    min_column, min_row, max_column, max_row = range_boundaries(range_ref)
    return Reference(
        worksheet,
        min_col=min_column,
        min_row=min_row,
        max_col=max_column,
        max_row=max_row,
    )


def add_chart(worksheet: Any, chart_spec: Mapping[str, Any]) -> None:
    chart_type = str(chart_spec.get("type", "column")).lower()
    if chart_type in ("bar", "column"):
        chart = BarChart()
        chart.type = "bar" if chart_type == "bar" else "col"
    elif chart_type == "line":
        chart = LineChart()
    elif chart_type == "pie":
        chart = PieChart()
    elif chart_type == "area":
        chart = AreaChart()
    else:
        raise SpreadsheetError("Unsupported chart type: " + chart_type)

    data_range = chart_spec.get("data")
    if not data_range:
        raise SpreadsheetError("chart.data is required")
    chart.add_data(
        reference_from_range(worksheet, str(data_range)),
        titles_from_data=bool(chart_spec.get("titles_from_data", True)),
        from_rows=bool(chart_spec.get("from_rows", False)),
    )
    if chart_spec.get("categories"):
        chart.set_categories(
            reference_from_range(worksheet, str(chart_spec["categories"]))
        )
    if chart_spec.get("title"):
        chart.title = str(chart_spec["title"])
    if chart_spec.get("style") is not None:
        chart.style = int(chart_spec["style"])
    chart.width = float(chart_spec.get("width", 12))
    chart.height = float(chart_spec.get("height", 7))
    if chart_spec.get("legend") is False:
        chart.legend = None
    if hasattr(chart, "x_axis") and chart_spec.get("x_axis_title"):
        chart.x_axis.title = str(chart_spec["x_axis_title"])
    if hasattr(chart, "y_axis") and chart_spec.get("y_axis_title"):
        chart.y_axis.title = str(chart_spec["y_axis_title"])
    worksheet.add_chart(chart, str(chart_spec.get("anchor", "F2")))


def apply_tables_and_charts(worksheet: Any, sheet_spec: Mapping[str, Any]) -> None:
    tables = sheet_spec.get("tables", [])
    if not isinstance(tables, list):
        raise SpreadsheetError("tables must be an array")
    for table_spec in tables:
        if not isinstance(table_spec, dict):
            raise SpreadsheetError("each tables entry must be an object")
        table_range = str(table_spec.get("range", "")).strip()
        table_name = str(table_spec.get("name", "")).strip()
        if not table_range or not table_name:
            raise SpreadsheetError("each table requires range and name")
        if not TABLE_NAME_RE.fullmatch(table_name):
            raise SpreadsheetError("Invalid Excel table name: " + table_name)
        table = Table(displayName=table_name, ref=table_range)
        table.tableStyleInfo = TableStyleInfo(
            name=str(table_spec.get("style", "TableStyleMedium2")),
            showFirstColumn=bool(table_spec.get("show_first_column", False)),
            showLastColumn=bool(table_spec.get("show_last_column", False)),
            showRowStripes=bool(table_spec.get("show_row_stripes", True)),
            showColumnStripes=bool(table_spec.get("show_column_stripes", False)),
        )
        worksheet.add_table(table)

    charts = sheet_spec.get("charts", [])
    if not isinstance(charts, list):
        raise SpreadsheetError("charts must be an array")
    for chart_spec in charts:
        if not isinstance(chart_spec, dict):
            raise SpreadsheetError("each charts entry must be an object")
        add_chart(worksheet, chart_spec)


def apply_sheet(
    worksheet: Any, sheet_spec: Mapping[str, Any], theme: Mapping[str, str]
) -> None:
    clear_ranges = sheet_spec.get("clear_ranges", [])
    if not isinstance(clear_ranges, list):
        raise SpreadsheetError("clear_ranges must be an array")
    for range_ref in clear_ranges:
        for cell in iter_cells_in_range(worksheet, str(range_ref)):
            cell.value = None

    if "rows" in sheet_spec:
        write_rows(
            worksheet,
            sheet_spec["rows"],
            positive_int(sheet_spec.get("start_row"), "start_row"),
            positive_int(sheet_spec.get("start_column"), "start_column"),
        )
    apply_cell_map(worksheet, sheet_spec.get("set_cells"))
    if "append_rows" in sheet_spec:
        write_rows(worksheet, sheet_spec["append_rows"], worksheet.max_row + 1, 1)

    if sheet_spec.get("freeze_panes"):
        worksheet.freeze_panes = str(sheet_spec["freeze_panes"])
    worksheet.sheet_view.showGridLines = bool(sheet_spec.get("show_grid_lines", True))
    if sheet_spec.get("tab_color"):
        worksheet.sheet_properties.tabColor = normalize_color(
            sheet_spec["tab_color"], "tab_color"
        )

    merged_cells = sheet_spec.get("merged_cells", [])
    if not isinstance(merged_cells, list):
        raise SpreadsheetError("merged_cells must be an array")
    for range_ref in merged_cells:
        normalized = str(range_ref)
        if normalized not in {str(item) for item in worksheet.merged_cells.ranges}:
            worksheet.merge_cells(normalized)

    apply_dimensions(worksheet, sheet_spec)
    apply_styles(worksheet, sheet_spec, theme)

    auto_filter = sheet_spec.get("auto_filter")
    if auto_filter is True:
        worksheet.auto_filter.ref = worksheet.dimensions
    elif isinstance(auto_filter, str) and auto_filter.strip():
        worksheet.auto_filter.ref = auto_filter.strip()

    apply_tables_and_charts(worksheet, sheet_spec)


def apply_metadata(workbook: Any, spec: Mapping[str, Any]) -> None:
    metadata = spec.get("metadata", {})
    if not isinstance(metadata, dict):
        raise SpreadsheetError("metadata must be an object")
    allowed = ("title", "subject", "creator", "description", "keywords", "category")
    for key in allowed:
        if key in metadata:
            setattr(workbook.properties, key, str(metadata[key]))
    if not workbook.properties.creator:
        workbook.properties.creator = "ArcForge"


def ensure_sheet_specs(spec: Mapping[str, Any]) -> List[Mapping[str, Any]]:
    sheets = spec.get("sheets")
    if not isinstance(sheets, list) or not sheets:
        raise SpreadsheetError("sheets must be a non-empty array")
    for sheet in sheets:
        if not isinstance(sheet, dict) or not str(sheet.get("name", "")).strip():
            raise SpreadsheetError("each sheet requires a name")
    return sheets


def create_workbook(spec: Mapping[str, Any]) -> Any:
    workbook = Workbook()
    workbook.remove(workbook.active)
    theme = merged_theme(spec)
    apply_metadata(workbook, spec)
    for sheet_spec in ensure_sheet_specs(spec):
        worksheet = workbook.create_sheet(str(sheet_spec["name"]))
        apply_sheet(worksheet, sheet_spec, theme)
    return workbook


def patch_workbook(input_path: Path, spec: Mapping[str, Any]) -> Any:
    if not input_path.is_file():
        raise SpreadsheetError("Input workbook does not exist: " + str(input_path))
    if input_path.suffix.lower() != ".xlsx":
        raise SpreadsheetError("Patch currently supports .xlsx input only")
    try:
        workbook = load_workbook(input_path, data_only=False, keep_links=True)
    except Exception as error:
        raise SpreadsheetError(
            "Failed to open input workbook: " + str(error)
        ) from error
    theme = merged_theme(spec)
    apply_metadata(workbook, spec)
    for sheet_spec in ensure_sheet_specs(spec):
        name = str(sheet_spec["name"])
        if name in workbook.sheetnames:
            worksheet = workbook[name]
        elif bool(sheet_spec.get("create_if_missing", False)):
            worksheet = workbook.create_sheet(name)
        else:
            raise SpreadsheetError(
                "Worksheet does not exist; set create_if_missing=true to add it: "
                + name
            )
        apply_sheet(worksheet, sheet_spec, theme)
    return workbook


def load_code_workbook(input_path: Optional[Path]) -> Any:
    if input_path is None:
        return Workbook()
    if not input_path.is_file():
        raise SpreadsheetError("Input workbook does not exist: " + str(input_path))
    if input_path.suffix.lower() != ".xlsx":
        raise SpreadsheetError("SpreadsheetCode currently supports .xlsx input only")
    try:
        return load_workbook(input_path, data_only=False, keep_links=True)
    except Exception as error:
        raise SpreadsheetError(
            "Failed to open input workbook: " + str(error)
        ) from error


def load_spreadsheet_code(script_path: Path) -> Tuple[Any, Dict[str, Any]]:
    if not script_path.is_file():
        raise SpreadsheetError("Spreadsheet code does not exist: " + str(script_path))
    if script_path.suffix.lower() != ".py":
        raise SpreadsheetError("Spreadsheet code path must end with .py")
    try:
        source_bytes = script_path.read_bytes()
    except OSError as error:
        raise SpreadsheetError("Failed to read spreadsheet code: " + str(error)) from error
    if len(source_bytes) > MAX_SCRIPT_BYTES:
        raise SpreadsheetError(
            f"Spreadsheet code exceeds the {MAX_SCRIPT_BYTES}-byte size limit"
        )
    try:
        source = source_bytes.decode("utf-8-sig")
    except UnicodeDecodeError as error:
        raise SpreadsheetError("Spreadsheet code must be UTF-8 text") from error
    try:
        tree = ast.parse(source, filename=str(script_path), mode="exec")
    except SyntaxError as error:
        location = f" at line {error.lineno}" if error.lineno else ""
        raise SpreadsheetError(
            "Spreadsheet code has invalid Python syntax" + location + ": " + error.msg
        ) from error
    validator = SpreadsheetCodeValidator()
    validator.visit(tree)
    compiled = compile(tree, str(script_path), "exec", dont_inherit=True, optimize=0)
    return compiled, {
        "path": str(script_path.resolve()),
        "sha256": hashlib.sha256(source_bytes).hexdigest(),
        "size_bytes": len(source_bytes),
        "ast_nodes": validator.node_count,
    }


def spreadsheet_code_environment(workbook: Any) -> Dict[str, Any]:
    return {
        "__builtins__": SAFE_CODE_BUILTINS,
        "workbook": workbook,
        "Alignment": Alignment,
        "AreaChart": AreaChart,
        "BarChart": BarChart,
        "Border": Border,
        "Comment": Comment,
        "Decimal": Decimal,
        "Font": Font,
        "GradientFill": GradientFill,
        "LineChart": LineChart,
        "NamedStyle": NamedStyle,
        "PatternFill": PatternFill,
        "PieChart": PieChart,
        "Protection": Protection,
        "Reference": Reference,
        "Side": Side,
        "Table": Table,
        "TableStyleInfo": TableStyleInfo,
        "copy": copy,
        "date": date,
        "datetime": datetime,
        "get_column_letter": get_column_letter,
        "timedelta": timedelta,
    }


def execute_spreadsheet_code(workbook: Any, script_path: Path) -> Dict[str, Any]:
    compiled, evidence = load_spreadsheet_code(script_path)
    environment = spreadsheet_code_environment(workbook)
    try:
        exec(compiled, environment, environment)
    except SpreadsheetError:
        raise
    except Exception as error:
        raise SpreadsheetError(
            f"Spreadsheet code failed with {type(error).__name__}: {error}"
        ) from error
    if not workbook.worksheets:
        raise SpreadsheetError("Spreadsheet code must leave at least one worksheet")
    if not any(worksheet.sheet_state == "visible" for worksheet in workbook.worksheets):
        raise SpreadsheetError("Spreadsheet code must leave at least one visible worksheet")
    return evidence


def normalized_output_path(path_value: str) -> Path:
    path = Path(path_value).expanduser().resolve()
    if path.suffix.lower() != ".xlsx":
        raise SpreadsheetError("Output path must end with .xlsx")
    return path


def atomic_save(workbook: Any, output_path: Path, force: bool) -> None:
    if output_path.exists() and not force:
        raise SpreadsheetError(
            "Output already exists. Use a new path or pass --force only after explicit approval: "
            + str(output_path)
        )
    output_path.parent.mkdir(parents=True, exist_ok=True)
    handle, temporary_name = tempfile.mkstemp(
        prefix=".arcforge-workbook-",
        suffix=".xlsx",
        dir=str(output_path.parent),
    )
    os.close(handle)
    temporary_path = Path(temporary_name)
    try:
        workbook.save(temporary_path)
        verification = load_workbook(temporary_path, read_only=True, data_only=False)
        verification.close()
        os.replace(temporary_path, output_path)
    except Exception:
        try:
            temporary_path.unlink(missing_ok=True)
        except OSError:
            pass
        raise


def inspect_workbook(path: Path) -> Dict[str, Any]:
    if not path.is_file():
        raise SpreadsheetError("Workbook does not exist: " + str(path))
    try:
        workbook = load_workbook(
            path, data_only=False, read_only=False, keep_links=True
        )
    except Exception as error:
        raise SpreadsheetError("Failed to inspect workbook: " + str(error)) from error

    sheets: List[Dict[str, Any]] = []
    total_formulas = 0
    total_nonempty = 0
    scan_truncated = False
    try:
        for worksheet in workbook.worksheets:
            formulas = 0
            nonempty = 0
            scanned = 0
            for row in worksheet.iter_rows():
                for cell in row:
                    scanned += 1
                    if scanned > MAX_INSPECTED_CELLS:
                        scan_truncated = True
                        break
                    if cell.value is not None:
                        nonempty += 1
                        if isinstance(cell.value, str) and cell.value.startswith("="):
                            formulas += 1
                if scanned > MAX_INSPECTED_CELLS:
                    break
            total_formulas += formulas
            total_nonempty += nonempty
            sheets.append(
                {
                    "name": worksheet.title,
                    "dimensions": worksheet.calculate_dimension(),
                    "max_row": worksheet.max_row,
                    "max_column": worksheet.max_column,
                    "nonempty_cells": nonempty,
                    "formula_cells": formulas,
                    "merged_ranges": [
                        str(item) for item in worksheet.merged_cells.ranges
                    ],
                    "tables": list(worksheet.tables.keys()),
                    "chart_count": len(worksheet._charts),
                    "freeze_panes": str(worksheet.freeze_panes)
                    if worksheet.freeze_panes
                    else None,
                    "auto_filter": worksheet.auto_filter.ref,
                }
            )
    finally:
        workbook.close()

    return {
        "path": str(path.resolve()),
        "size_bytes": path.stat().st_size,
        "sheet_count": len(sheets),
        "sheets": sheets,
        "total_nonempty_cells": total_nonempty,
        "total_formula_cells": total_formulas,
        "scan_truncated": scan_truncated,
        "formula_results_calculated": False,
    }


def run_create(args: argparse.Namespace) -> Dict[str, Any]:
    spec, _ = load_json_object(args.spec)
    output_path = normalized_output_path(args.output)
    workbook = create_workbook(spec)
    atomic_save(workbook, output_path, args.force)
    workbook.close()
    return {"action": "created", "workbook": inspect_workbook(output_path)}


def run_patch(args: argparse.Namespace) -> Dict[str, Any]:
    spec, _ = load_json_object(args.spec)
    input_path = Path(args.input).expanduser().resolve()
    output_path = normalized_output_path(args.output)
    workbook = patch_workbook(input_path, spec)
    atomic_save(workbook, output_path, args.force)
    workbook.close()
    return {
        "action": "patched",
        "input": str(input_path),
        "workbook": inspect_workbook(output_path),
    }


def run_code(args: argparse.Namespace) -> Dict[str, Any]:
    input_path = (
        Path(args.input).expanduser().resolve() if args.input is not None else None
    )
    script_path = Path(args.script).expanduser().resolve()
    output_path = normalized_output_path(args.output)
    workbook = load_code_workbook(input_path)
    try:
        code_evidence = execute_spreadsheet_code(workbook, script_path)
        atomic_save(workbook, output_path, args.force)
    finally:
        workbook.close()
    return {
        "action": "code_executed",
        "input": str(input_path) if input_path is not None else None,
        "script": code_evidence,
        "workbook": inspect_workbook(output_path),
    }


def run_inspect(args: argparse.Namespace) -> Dict[str, Any]:
    return {
        "action": "inspected",
        "workbook": inspect_workbook(Path(args.input).expanduser().resolve()),
    }


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        description="Create, patch, code, and structurally inspect XLSX workbooks for ArcForge."
    )
    subparsers = parser.add_subparsers(dest="command", required=True)

    create_parser = subparsers.add_parser(
        "create", help="Create a new XLSX workbook from JSON"
    )
    create_parser.add_argument("--spec", required=True, help="UTF-8 JSON specification")
    create_parser.add_argument("--output", required=True, help="Destination .xlsx path")
    create_parser.add_argument(
        "--force", action="store_true", help="Overwrite the exact output path"
    )
    create_parser.set_defaults(handler=run_create)

    patch_parser = subparsers.add_parser(
        "patch", help="Patch an XLSX workbook from JSON"
    )
    patch_parser.add_argument("--input", required=True, help="Existing .xlsx path")
    patch_parser.add_argument(
        "--spec", required=True, help="UTF-8 JSON patch specification"
    )
    patch_parser.add_argument("--output", required=True, help="Destination .xlsx path")
    patch_parser.add_argument(
        "--force", action="store_true", help="Overwrite the exact output path"
    )
    patch_parser.set_defaults(handler=run_patch)

    code_parser = subparsers.add_parser(
        "code", help="Run validated workbook-only Python against an XLSX workbook"
    )
    code_parser.add_argument(
        "--input", help="Optional existing .xlsx path; omit to start a new workbook"
    )
    code_parser.add_argument(
        "--script", required=True, help="UTF-8 workbook-only Python script"
    )
    code_parser.add_argument("--output", required=True, help="Destination .xlsx path")
    code_parser.add_argument(
        "--force", action="store_true", help="Overwrite the exact output path"
    )
    code_parser.set_defaults(handler=run_code)

    inspect_parser = subparsers.add_parser(
        "inspect", help="Print a structural workbook summary"
    )
    inspect_parser.add_argument("--input", required=True, help="Existing .xlsx path")
    inspect_parser.set_defaults(handler=run_inspect)
    return parser


def main(argv: Optional[Sequence[str]] = None) -> int:
    parser = build_parser()
    args = parser.parse_args(argv)
    try:
        require_openpyxl()
        result = args.handler(args)
        print(json.dumps(result, ensure_ascii=False, indent=2))
        return 0
    except SpreadsheetError as error:
        print("error: " + str(error), file=sys.stderr)
        return 2
    except Exception as error:
        print("error: unexpected spreadsheet failure: " + str(error), file=sys.stderr)
        return 3


if __name__ == "__main__":
    raise SystemExit(main())
